import flet as ft
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

# Arquivo original de referência
arquivo_original = 'prototipo.xlsx'

# Função para aplicar estilo à célula (bordas, centralização, negrito e fundo cinza)
def aplicar_estilo(celula, total=False):
    """Aplica bordas, centralização e, opcionalmente, o estilo de total (negrito e fundo cinza)."""
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    alinhamento = Alignment(horizontal="center", vertical="center")
    celula.border = borda
    celula.alignment = alinhamento
    
    if total:
        # Aplicando negrito e preenchimento cinza
        celula.font = Font(bold=True)
        celula.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

def main(page):
    # Configurações iniciais da janela
    page.title = "Gestor de Itens - Excel"
    page.window_width = 350
    page.window_height = 700

    # Variáveis globais
    linha_inicial = 4  # Começa em 4 porque a partir de A4 serão preenchidos os itens
    itens_adicionados = []  # Lista para armazenar os itens adicionados


    # Carrega o arquivo Excel para trabalhar com ele
    workbook = load_workbook('prototipo.xlsx')
    sheet = workbook.active
    

    def adicionar_item(event):
        """Adiciona item, quantidade e potência à lista."""
        # Descobrir a próxima linha disponível antes das duas últimas linhas em branco
        linha_atual = linha_inicial
        while sheet[f"A{linha_atual}"].value is not None:
            linha_atual += 1

        item = submenu_dropdown.value if submenu_dropdown.visible else item_dropdown.value
        quantidade = quantidade_input.value

        # Mapeamento de potência dos itens
        potencias = {
            "SPLIT 9.000BTUS": "0.99",
            "SPLIT 12.000BTUS": "1.26",
            "SPLIT 18.000BTUS": "2.18",
            "SPLIT 22.000BTUS": "2.43",
            "SPLIT 24.000BTUS": "2.89",
            "ILUMINAÇÃO": "0.05",
            "TOMADA": "0.10",
            "MOTOR DE PORTÃO": "0.40",
            "BOMBA 1/4 CV": "0.34",
            "BOMBA 1/2 CV": "0.61",
            "BOMBA 3/4 CV": "0.85",
            "BOMBA 1 CV": "1.05",
            "BOMBA 2 CV": "1.47",
            "BOMBA 3 CV": "2.20",
            "CHUVEIRO ELÉTRICO": "4.5" 

        }

        potencia = potencias.get(item, "0")

        # Validação dos campos
        if not item:
            feedback_text.value = "Selecione um item antes de adicionar."
            feedback_text.color = ft.Colors.RED
            page.update()
            return

        if not quantidade or not quantidade.isdigit():
            feedback_text.value = "Digite uma quantidade válida."
            feedback_text.color = ft.Colors.RED
            page.update()
            return

        # Verifica se o item já foi adicionado
        for item_existente in itens_adicionados:
            if item_existente['item'] == item:
                item_existente['quantidade'] += int(quantidade)  # Soma a quantidade
                feedback_text.value = "Quantidade atualizada!"
                feedback_text.color = ft.Colors.GREEN
                page.update()
                return

        # Adiciona o item à lista de itens
        itens_adicionados.append({
            'item': item,
            'quantidade': int(quantidade),
            'potencia': float(potencia)  # A potência agora será um valor numérico, inclusive para o SPLIT 12.000
        })

        # Atualiza a lista deslizante
        if len(itens_column.controls) >= 4:
            itens_column.controls.pop(0)
        itens_column.controls.append(ft.Checkbox(label=f"Item: {item}, Quantidade: {quantidade}"))
        
        # Feedback
        feedback_text.value = "Item adicionado à lista!"
        feedback_text.color = ft.colors.GREEN

        # Limpa campos
        quantidade_input.value = ""
        item_dropdown.value = ""
        submenu_dropdown.value = ""
        page.update()

    

    def salvar_arquivo(event):
        
        """Atualiza o arquivo Excel com os itens da lista e salva o arquivo."""
        # Atualiza o arquivo Excel com os itens da lista
        linha_atual = linha_inicial
        for item in itens_adicionados:
            sheet[f"A{linha_atual}"] = item['item']
            sheet[f"B{linha_atual}"] = item['quantidade']
            sheet[f"C{linha_atual}"] = item['potencia']
            sheet[f"D{linha_atual}"] = f"=B{linha_atual}*C{linha_atual}"
            sheet[f"F{linha_atual}"] = f"=D{linha_atual}/E{linha_atual}"
            sheet[f"H{linha_atual}"] = f"=D{linha_atual}*G{linha_atual}"
            sheet[f"I{linha_atual}"] = f"=F{linha_atual}*G{linha_atual}"


            # Lógica para coluna E (somente 0.90 para os SPLIT)
            if "SPLIT" in item['item']:
                sheet[f"E{linha_atual}"] = 0.90  # Atribui 0.90 para qualquer item SPLIT
            if "ILUMINAÇÃO" in item['item']:
                sheet[f"E{linha_atual}"] = 0.82 
            if "TOMADA" in item['item']:
                sheet[f"E{linha_atual}"] = 0.80
            if "MOTOR DE PORTÃO" in item['item']:
                sheet[f"E{linha_atual}"] = 0.80
            if "BOMBA 1/4 CV" in item['item']:
                sheet[f"E{linha_atual}"] = 0.80
            if "BOMBA 1/2 CV" in item['item']:
                sheet[f"E{linha_atual}"] = 0.80
            if "BOMBA 3/4 CV" in item['item']:
                sheet[f"E{linha_atual}"] = 0.80
            if "BOMBA 1 CV" in item['item']:
                sheet[f"E{linha_atual}"] = 0.80
            if "BOMBA 2 CV" in item['item']:
                sheet[f"E{linha_atual}"] = 0.80
            if "BOMBA 3 CV" in item['item']:
                sheet[f"E{linha_atual}"] = 0.80
            if "CHUVEIRO ELÉTRICO" in item['item']:
                sheet[f"E{linha_atual}"] = 1.00
            
            # Lógica para coluna G (desconto por quantidade de SPLIT)
            if "SPLIT" in item['item']:
                quantidade_splits = item['quantidade']
                if 1 <= quantidade_splits <= 2:
                    sheet[f"G{linha_atual}"] = 1.00
                elif quantidade_splits == 3:
                    sheet[f"G{linha_atual}"] = 0.88
                elif quantidade_splits == 4:
                    sheet[f"G{linha_atual}"] = 0.82
                elif quantidade_splits == 5:
                    sheet[f"G{linha_atual}"] = 0.78
                elif quantidade_splits == 6:
                    sheet[f"G{linha_atual}"] = 0.76
                elif quantidade_splits == 7:
                    sheet[f"G{linha_atual}"] = 0.74
                elif quantidade_splits == 8:
                    sheet[f"G{linha_atual}"] = 0.72
                elif quantidade_splits >= 10:
                    sheet[f"G{linha_atual}"] = 0.70
            
            if "ILUMINAÇÃO" in item['item']:
                sheet[f"G{linha_atual}"] = 0.70
            if "TOMADA" in item['item']:
                sheet[f"G{linha_atual}"] = 0.45
            if "MOTOR DE PORTÃO" in item['item']:
                sheet[f"G{linha_atual}"] = 0.40
            if "CHUVEIRO ELÉTRICO" in item['item']:
                sheet[f"G{linha_atual}"] = 0.45
            if "BOMBA 1/4 CV" in item['item']:
                sheet[f"G{linha_atual}"] = 0.40
            if "BOMBA 1/2 CV" in item['item']:
                sheet[f"G{linha_atual}"] = 0.40
            if "BOMBA 3/4 CV" in item['item']:
                sheet[f"G{linha_atual}"] = 0.40
            if "BOMBA 1 CV" in item['item']:
                sheet[f"G{linha_atual}"] = 0.40
            if "BOMBA 2 CV" in item['item']:
                sheet[f"G{linha_atual}"] = 0.40
            if "BOMBA 3 CV" in item['item']:
                sheet[f"G{linha_atual}"] = 0.40

            # Aplica bordas e centralização em todas as células da linha (A até G)
            for coluna in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                aplicar_estilo(sheet[f"{coluna}{linha_atual}"])
            
            # Aplicar número formatado para 2 casas decimais
            for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
                sheet[f"{col}{linha_atual}"].number_format = '0.00'

            linha_atual += 1

        # Agora a última linha é definida corretamente após o loop
        ultima_linha = linha_atual

        # Adiciona a linha de soma ao final da tabela
        sheet[f"A{ultima_linha}"] = ""
        sheet[f"D{ultima_linha}"] = f"=SUM(D{linha_inicial}:D{ultima_linha-1})"
        sheet[f"F{ultima_linha}"] = f"=SUM(F{linha_inicial}:F{ultima_linha-1})"
        sheet[f"H{ultima_linha}"] = f"=SUM(H{linha_inicial}:H{ultima_linha-1})"
        sheet[f"I{ultima_linha}"] = f"=SUM(I{linha_inicial}:I{ultima_linha-1})"
        sheet[f"K3"] = f"=IF(D{ultima_linha}<=12, \"MOFÁSICO\", \"TRIFÁSICO\")"
        
        sheet[f"K4"] = f"=IF(D{ultima_linha}<=4, \"CABO 4mm²\",IF(D{ultima_linha}<=8, \"CABO 6mm²\", IF(D{ultima_linha}<=12, \"CABO 10mm²\", IF(D{ultima_linha}<=20, \"CABO 6mm²\",IF(D{ultima_linha}<=30, \"CABO 10mm²\", IF(D{ultima_linha}<=40, \"CABO 16mm²\"))))))"
        sheet[f"K5"] = f"=IF(D{ultima_linha}<=4, \"DISJUNTOR 25A MONOFASICO\",IF(D{ultima_linha}<=8, \"DISJUNTOR 40A MONOFASICO\", IF(D{ultima_linha}<=12, \"DISJUNTOR 63A MONOFASICO\", IF(D{ultima_linha}<=20, \"DISJUNTOR 40A TRIFASICO\",IF(D{ultima_linha}<=30, \"DISJUNTOR 63A TRIFASICO\", IF(D{ultima_linha}<=40, \"DISJUNTOR 80A TRIFASICO\"))))))"

        #sheet[f"K4"] = f"=IF(D{ultima_linha}<=20, \"CABO 6mm²\",IF(D{ultima_linha}<=30, \"CABO 10mm²\", IF(D{ultima_linha}<=40, \"CABO 16mm²\")))"
        #sheet[f"K5"] = f"=IF(D{ultima_linha}<=20, \"DISJUNTOR 40A TRIFASICO\",IF(D{ultima_linha}<=30, \"DISJUNTOR 63A TRIFASICO\", IF(D{ultima_linha}<=40, \"DISJUNTOR 80A TRIFASICO\")))"

        # Adiciona título na coluna K (Tipo de Ligação)
        sheet["K2"] = "TIPO DE LIGAÇÃO"
        aplicar_estilo(sheet["K2"], total=True)
        aplicar_estilo(sheet["K3"], total=False)
        aplicar_estilo(sheet["K4"], total=False)
        aplicar_estilo(sheet["K5"], total=False)
    
        # Aplica bordas e centralização na linha de total (somente coluna D)
        aplicar_estilo(sheet[f"D{ultima_linha}"], total=True)
        aplicar_estilo(sheet[f"F{ultima_linha}"], total=True)
        aplicar_estilo(sheet[f"H{ultima_linha}"], total=True)
        aplicar_estilo(sheet[f"I{ultima_linha}"], total=True)

        # Aplica bordas na última linha (TOTAL) também
        for coluna in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:  # Garantindo que todas as colunas da última linha tenham borda
            aplicar_estilo(sheet[f"{coluna}{ultima_linha}"])
         
        # Formatação na linha de totais
        for col in ['D', 'F', 'H', 'I']:
            sheet[f"{col}{linha_atual}"].number_format = '0.00'
            aplicar_estilo(sheet[f"{col}{linha_atual}"], total=True)

        
        try:
            fnalizar_button.visible = False
            if nome_input.value.strip() == "":
                feedback_text.value = "O nome precisa ser digitado."
                feedback_text.color = ft.Colors.RED
            else:
                workbook.save(f'Quadro de Cargas {nome_input.value}.xlsx')
                feedback_text.value = f'Quadro de cargas salvo com sucesso'
                feedback_text.color = ft.Colors.GREEN

            
            
        
        except Exception as e:
            feedback_text.value = f"Erro ao salvar o arquivo: {e}", 
            feedback_text.color = ft.Colors.RED
            
        page.update()

        #troca a visibilidade do botão salvar e caixa de texto de nome
    def chama_botão_salvar(event):
        nome_input.visible = True
        salvar_button.visible = True
        feedback_text.value = ''
        page.update()
        



    def atualizar_submenu(event):
        """Atualiza o submenu com base na seleção principal."""
        if item_dropdown.value == "SPLITS":
            submenu_dropdown.visible = True
            submenu_dropdown.options = [
                ft.dropdown.Option("SPLIT 9.000BTUS"),
                ft.dropdown.Option("SPLIT 12.000BTUS"),
                ft.dropdown.Option("SPLIT 18.000BTUS"),
                ft.dropdown.Option("SPLIT 22.000BTUS"),
                ft.dropdown.Option("SPLIT 24.000BTUS")
            ]
        elif item_dropdown.value == "BOMBAS":
            submenu_dropdown.visible = True
            submenu_dropdown.options = [
                ft.dropdown.Option("BOMBA 1/4 CV"),
                ft.dropdown.Option("BOMBA 1/2 CV"),
                ft.dropdown.Option("BOMBA 3/4 CV"),
                ft.dropdown.Option("BOMBA 1 CV"),
                ft.dropdown.Option("BOMBA 2 CV"),
                ft.dropdown.Option("BOMBA 3 CV"),
            ]
        else:
            submenu_dropdown.visible = False
        page.update()

    # Elementos da interface

    # Feedback para o usuário
    feedback_text = ft.Text(color="green")

    # Dropdown principal
    item_dropdown = ft.Dropdown(
        width=300,
        label="Selecione o item",
        options=[
            ft.dropdown.Option("SPLITS"),
            ft.dropdown.Option("ILUMINAÇÃO"),
            ft.dropdown.Option("TOMADA"),
            ft.dropdown.Option("MOTOR DE PORTÃO"),
            ft.dropdown.Option("BOMBAS"),
            ft.dropdown.Option("CHUVEIRO ELÉTRICO")
        ],
        on_change=atualizar_submenu
    )

    # Submenu (para SPLITS e BOMBA)
    submenu_dropdown = ft.Dropdown(
        width=300,
        label="Selecione o tipo",
        options=[],
        visible=False
    )

    # Campo para quantidade
    quantidade_input = ft.TextField(label="Quantidade")

    nome_input = ft.TextField(label="Digite o nome do cliente",visible=False)

    # Botão para adicionar item
    adicionar_button = ft.ElevatedButton(
        text="Adicionar",
        on_click=adicionar_item
    )

    # Botão para finalizar itens
    fnalizar_button = ft.ElevatedButton(
        text="Finalizar",
        on_click= chama_botão_salvar

    )

     # Botão para salvar arquivo
    salvar_button = ft.ElevatedButton(
        text="Salvar Arquivo",
        on_click= salvar_arquivo,
        visible= False
    )

    # Adicionar o botão em uma Row para centralizar
    finalizar_button_container = ft.Row(
        controls=[fnalizar_button],
        alignment=ft.MainAxisAlignment.CENTER)

    # Adicionar o botão em uma Row para centralizar
    adicionar_button_container = ft.Row(
        controls=[adicionar_button],
        alignment=ft.MainAxisAlignment.CENTER)

    # Adicionar o botão em uma Row para centralizar
    salvar_button_container = ft.Row(
        controls=[salvar_button],
        alignment=ft.MainAxisAlignment.CENTER)

    # Coluna para exibir os itens adicionados com rolagem
    itens_column = ft.ListView(
        spacing=10,
        padding=10,
        first_item_prototype= True,
        on_scroll= True, # Rola automaticamente para o final quando um novo item é adicionado
        height=200  # Limita a altura visível da lista
    )

    # Layout principal
    page.add(
        ft.Container(
            width=350,
            height=60,
            content=ft.Row(
                controls=[ft.Text("Gestor de Cargas", size=20)],
                alignment=ft.MainAxisAlignment.CENTER
            ),
            bgcolor="orange"
        ),
        item_dropdown,
        submenu_dropdown,
        quantidade_input,
        adicionar_button_container,
        ft.Container(content=itens_column, height=200, bgcolor=ft.colors.GREY),
        finalizar_button_container,
        nome_input,
        salvar_button_container,
        
        feedback_text
    )

ft.app(target=main)
