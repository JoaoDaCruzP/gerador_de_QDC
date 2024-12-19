import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

# Criação de um DataFrame com a data e hora atuais
data_hora_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
df = pd.DataFrame({"Data e Hora": [data_hora_atual]})

# Salvar o DataFrame em um arquivo Excel
caminho_arquivo = 'data_hora_atual.xlsx'
df.to_excel(caminho_arquivo, index=False, engine='openpyxl')

# Abrir o arquivo Excel com openpyxl para formatar
wb = openpyxl.load_workbook(caminho_arquivo)
ws = wb.active

# Definir estilo de negrito e fundo amarelo
celula = ws['A2']
celula.font = Font(bold=True)
celula.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Ajustar a largura da coluna com base no tamanho do conteúdo
coluna = 'A'
max_len = 0
for cell in ws[coluna]:
    try:
        if len(str(cell.value)) > max_len:
            max_len = len(cell.value)
    except:
        pass
# Definir a largura da coluna para se ajustar ao conteúdo
ws.column_dimensions[coluna].width = max_len + 2  # +2 para dar uma margem extra

# Salvar as alterações
wb.save(caminho_arquivo)

print(f"Arquivo gerado com sucesso: {caminho_arquivo}")
